import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter

def analyze_excel_layout(filepath):
    """Excelファイルの構造を分析"""
    try:
        # openpyxlでワークブックを開く
        workbook = openpyxl.load_workbook(filepath)
        
        print(f"ワークシート一覧: {workbook.sheetnames}")
        
        for sheet_name in workbook.sheetnames:
            print(f"\n=== シート: {sheet_name} ===")
            sheet = workbook[sheet_name]
            
            # シートの範囲を取得
            max_row = sheet.max_row
            max_col = sheet.max_column
            print(f"データ範囲: A1:{get_column_letter(max_col)}{max_row}")
            
            # セルの内容を出力（最初の50行まで）
            for row_num in range(1, min(51, max_row + 1)):
                row_data = []
                for col_num in range(1, max_col + 1):
                    cell = sheet.cell(row=row_num, column=col_num)
                    value = cell.value
                    if value is not None:
                        # セルの背景色やボーダーの情報も取得
                        fill = cell.fill
                        border = cell.border
                        font = cell.font
                        
                        cell_info = {
                            'value': str(value),
                            'coordinate': cell.coordinate,
                            'bg_color': fill.start_color.rgb if fill.start_color.rgb != '00000000' else None,
                            'font_bold': font.bold if font.bold else False,
                            'merged': False
                        }
                        
                        # マージされたセルかチェック
                        for merged_range in sheet.merged_cells.ranges:
                            if cell.coordinate in merged_range:
                                cell_info['merged'] = True
                                cell_info['merged_range'] = str(merged_range)
                                break
                        
                        row_data.append(cell_info)
                
                if row_data:
                    print(f"行 {row_num}:")
                    for cell_info in row_data:
                        coord = cell_info['coordinate']
                        value = cell_info['value']
                        extras = []
                        if cell_info['bg_color']:
                            extras.append(f"背景:{cell_info['bg_color']}")
                        if cell_info['font_bold']:
                            extras.append("太字")
                        if cell_info['merged']:
                            extras.append(f"結合:{cell_info['merged_range']}")
                        
                        extra_str = f" [{', '.join(extras)}]" if extras else ""
                        print(f"  {coord}: {value}{extra_str}")
                    print()
    
    except Exception as e:
        print(f"エラー: {e}")
        
        # 代替方法でpandasを使用
        try:
            print("\n=== pandasでの読み取り ===")
            df = pd.read_excel(filepath, sheet_name=None)
            
            for sheet_name, data in df.items():
                print(f"\nシート: {sheet_name}")
                print(f"形状: {data.shape}")
                print("データの内容:")
                print(data.to_string())
                print("\n" + "="*50)
        
        except Exception as e2:
            print(f"pandas読み取りエラー: {e2}")

if __name__ == "__main__":
    filepath = "/mnt/c/ClaudeCode/APPS/ELIMS/screen_images/登録画面.xlsx"
    analyze_excel_layout(filepath)